from pathlib import Path

from numpy import ndarray
import pandas as pd
import MDAnalysis as mda
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from ..constants import Constants
from ..logger import Logger

from .path_builder import PathBuilder
from .pdb_file_builder import PdbFileBuilder
from .file_manager import FileManager


logger = Logger("FileWriter")


class FileWriter(FileManager):
    # dat_file_first_lines: str = "        210         150           3  1.000000000000000E-002       10000\n" + \
    #                             "   5.00000000000000             3023   1.00000000000000\n"

    @classmethod
    def write_dat_file(
        cls,
        data_lines: list[str] | ndarray,
        path_to_file: Path | None = None,
        structure_folder: str | None = None,
        overwrite: bool = True,
        filename: str = Constants.filenames.INIT_DAT_FILE,
    ) -> None:
        """For the path you can provide either path_to_file or structure_folder."""

        # TODO: refactor to use DataConverter.convert_ndarray_to_dat
        try:
            if len(data_lines) == 0:
                logger.warning("No data for .dat file.")
                return

            if path_to_file is None:
                if structure_folder is None:
                    raise ValueError(
                        "Provide either path_to_file or structure_folder param for FileWriter.write_dat_file")

                path_to_file = PathBuilder.build_path_to_result_data_file(
                    structure_folder,
                    file=filename)

            if overwrite is False and path_to_file.exists():
                # Don't overwrite existing file
                return

            # Convert ndarray to list[str]
            if isinstance(data_lines, ndarray):
                data_lines = [f"{i[0]}\t{i[1]}\t{i[2]}" for i in data_lines]

            with Path(path_to_file).open("w") as dat_file:
                # dat_file.write(cls.dat_file_first_lines)

                for line in data_lines:
                    if "\n" not in line:
                        line += "\n"
                    dat_file.write(line)

            logger.info(f"File saved: {path_to_file}")

        except Exception as e:
            logger.error(f".dat file not saved: {e}")

    @staticmethod
    def write_pdb_file(
        data_lines: list[str] | ndarray,
        path_to_file: Path | None = None,
        structure_folder: str | None = None,
        overwrite: bool = True,
    ) -> None:
        """For the path you can provide either path_to_file or structure_folder."""

        # TODO: refactor to use DataConverter.convert_ndarray_to_pdb
        try:
            if len(data_lines) == 0:
                logger.warning("No data for .dat file.")
                return

            if path_to_file is None:
                if structure_folder is None:
                    raise ValueError(
                        "Provide either path_to_file or structure_folder param for FileWriter.write_pdb_file")

                path_to_file = PathBuilder.build_path_to_result_data_file(
                    structure_folder,
                    file=Constants.filenames.INIT_DAT_FILE)

            if overwrite is False and path_to_file.exists():
                # Don't overwrite existing file
                return

            # Convert ndarray to list[str]
            if isinstance(data_lines, ndarray):
                data_lines = [
                    PdbFileBuilder.build_pdb_line(
                        coords=[str(coord[0]), str(coord[1]), str(coord[2])], atom_id=i
                    ) for i, coord in enumerate(data_lines)
                ]

            with Path(path_to_file).open("w") as pdb_file:
                # Write the PDB file header
                pdb_file.write(PdbFileBuilder.first_lines)

                # Write the data
                for line in data_lines:
                    pdb_file.write(line)

                # Write the PDB file footer
                pdb_file.write(PdbFileBuilder.get_end_lines(num_of_lines=len(data_lines)))

        except Exception as e:
            logger.error(f".pdb file not saved: {e}")

    @classmethod
    def write_excel_file(
            cls,
            df: pd.DataFrame,
            structure_folder: str,
            sheet_name: str,
            file_name: str,
            folder_path: Path | str | None = None,
            is_init_data_dir: bool = True,
    ) -> Path | None:
        """
        Write a pandas DataFrame to an Excel file.

        Parameters:
        - df: pd.DataFrame, the data to write to the Excel file.
        - structure_folder: str, the name of the structure folder.
        - folder_path: Path | str | None, the base folder path. If None, uses the default logic from PathBuilder.
        - file_name: str, the Excel file name to write.
        - sheet_name: str, the name of the sheet where data will be written.
        - is_init_data_dir: bool | None: to build path to the specific dir. If it's False - builds path to result data.
        """

        path_to_file: Path = cls._get_path_to_file(structure_folder, file_name, folder_path, is_init_data_dir)

        # Ensure the directory exists
        path_to_file.parent.mkdir(parents=True, exist_ok=True)

        try:
            if isinstance(df.columns, pd.MultiIndex):
                # Handle MultiIndex columns
                with pd.ExcelWriter(path_to_file, engine='openpyxl') as writer:
                    # Write the DataFrame to the Excel file
                    df.to_excel(writer, sheet_name=sheet_name, startrow=1, header=False, index=True)

                    # Access the workbook and the sheet
                    worksheet: Worksheet = writer.sheets[sheet_name]

                    # Write the MultiIndex header
                    for idx, col in enumerate(df.columns):
                        worksheet.cell(row=1, column=idx + 2, value=col[0])
                        worksheet.cell(row=2, column=idx + 2, value=col[1])

                    # Merge cells for the top-level headers
                    for col in df.columns.levels[0]:
                        col_indices: list[int] = [i for i, c in enumerate(df.columns) if c[0] == col]
                        if col_indices:
                            worksheet.merge_cells(
                                start_row=1, start_column=col_indices[0] + 2,
                                end_row=1, end_column=col_indices[-1] + 2,
                            )
                            # Center align the merged cells
                            worksheet.cell(row=1, column=col_indices[0] + 2).alignment = Alignment(horizontal='center')
            else:
                # Write the DataFrame to an Excel file
                df.to_excel(path_to_file, sheet_name=sheet_name, index=False, engine="openpyxl")

            logger.info(f"Data successfully written to {path_to_file}")
            return path_to_file

        except Exception as e:
            logger.error(f"Failed to write file {path_to_file}: {e}")

    @staticmethod
    def write_pdb_from_mda(output_pdb_file: Path, atoms) -> None:
        with mda.Writer(output_pdb_file, n_atoms=atoms.n_atoms) as PDB:
            PDB.write(atoms)
